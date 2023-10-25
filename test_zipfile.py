import zipfile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from PIL import Image




def test_pdf():
    with zipfile.ZipFile('tmp/arhiv.zip') as zipf:
        with zipf.open('color_map.pdf') as pdffile:
            pdffile = PdfReader(pdffile)
            page = pdffile.pages[1]
            text = page.extract_text()
            assert text in 'Teinture Francaise'

            number_of_page = len(pdffile.pages)
            assert number_of_page == 5



def test_txt():
    with zipfile.ZipFile('tmp/arhiv.zip') as zipf:
        with zipf.open('print.txt') as txtf:
            s = txtf.read()
            assert s in b'test txt file'


def test_xlsx():
    with zipfile.ZipFile('tmp/arhiv.zip') as zipf:
        with zipf.open('exceltest.xlsx') as xlsxf:
            xlsx_file = load_workbook(xlsxf)
            sheet = xlsx_file.active
            value = sheet.cell(row=7, column=4).value
            assert value == 14


def test_foto():
    with zipfile.ZipFile('tmp/arhiv.zip') as zipf:
        with zipf.open('foto.jpg') as fotofile:
            foto = Image.open(fotofile)
            f=foto.format.format(foto)
            assert f in 'JPEG'

            s=foto.size
            str_s=str(s)
            assert str_s in '(1920, 1080)'
